#!/usr/bin/env python3
"""
Script per generare report direzionale completo
Report Excel con executive summary, analisi PNRR, e dettagli per struttura

Utilizzo:
    python genera_report_direzione.py

Output:
    report_direzione_telemedicina_YYYYMMDD.xlsx
"""

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def crea_report_direzione():
    """Genera report Excel completo per la direzione"""

    # Carica dati
    print("📊 Caricamento dati...")
    df_strutture = pd.read_csv('strutture_sanitarie.csv')
    df_catalogo = pd.read_csv('dotazioni_telemedicina_catalogo.csv')
    df_dotazioni = pd.read_csv('dotazioni_strutture_telemedicina.csv')

    # Merge per analisi
    df_merge = df_dotazioni.merge(
        df_catalogo,
        left_on='Codice_Dotazione',
        right_on='Codice',
        how='left'
    )

    df_merge['Quantita_Da_Acquistare'] = df_merge['Quantita_Richiesta'] - df_merge['Quantita_Presente']
    df_merge['Quantita_Da_Acquistare'] = df_merge['Quantita_Da_Acquistare'].clip(lower=0)
    df_merge['Costo_Totale'] = df_merge['Quantita_Da_Acquistare'] * df_merge['Costo_Unitario_EUR']

    # Merge con strutture
    df_merge = df_merge.merge(
        df_strutture[['Codice', 'Nome_Struttura', 'Tipologia', 'Zona', 'PNRR']],
        left_on='Codice_Struttura',
        right_on='Codice',
        how='left',
        suffixes=('', '_Strutt')
    )

    # Nome file output
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"report_direzione_telemedicina_{timestamp}.xlsx"

    print(f"📝 Generazione report: {filename}")

    # 1. EXECUTIVE SUMMARY
    print("  → Executive Summary")
    summary_data = {
        'Indicatore': [
            'Strutture Totali',
            '  - Case di Comunità (CDC)',
            '  - Ospedali di Comunità (ODC)',
            '',
            'Strutture PNRR',
            'Strutture non-PNRR',
            '',
            'FABBISOGNO DA FINANZIARE',
            '  - Interventi PNRR (priorità)',
            '  - Interventi non-PNRR',
            '',
            'Budget già finanziato',
            'Dotazioni già presenti (valore)',
            '',
            'TOTALE INVESTIMENTO NECESSARIO'
        ],
        'Valore': [],
        'Note': []
    }

    n_strutture = len(df_strutture)
    n_cdc = len(df_strutture[df_strutture['Tipologia'] == 'CdC'])
    n_odc = len(df_strutture[df_strutture['Tipologia'] == 'OdC'])
    n_pnrr = len(df_strutture[df_strutture['PNRR'] == 'SI'])
    n_non_pnrr = len(df_strutture[df_strutture['PNRR'] == 'NO'])

    # Calcoli finanziamento
    df_da_acq = df_merge[df_merge['Stato_Finanziamento'] == 'DA_ACQUISTARE']
    df_finanz = df_merge[df_merge['Stato_Finanziamento'] == 'FINANZIATO']
    df_presente = df_merge[df_merge['Stato_Finanziamento'] == 'PRESENTE']

    costo_da_acq_pnrr = df_da_acq[df_da_acq['PNRR'] == 'SI']['Costo_Totale'].sum()
    costo_da_acq_no = df_da_acq[df_da_acq['PNRR'] == 'NO']['Costo_Totale'].sum()
    costo_finanz = (df_finanz['Quantita_Richiesta'] * df_finanz['Costo_Unitario_EUR']).sum()
    costo_presente = (df_presente['Quantita_Presente'] * df_presente['Costo_Unitario_EUR']).sum()

    summary_data['Valore'] = [
        n_strutture,
        n_cdc,
        n_odc,
        '',
        n_pnrr,
        n_non_pnrr,
        '',
        f"€ {costo_da_acq_pnrr + costo_da_acq_no:,.2f}",
        f"€ {costo_da_acq_pnrr:,.2f}",
        f"€ {costo_da_acq_no:,.2f}",
        '',
        f"€ {costo_finanz:,.2f}",
        f"€ {costo_presente:,.2f}",
        '',
        f"€ {costo_da_acq_pnrr + costo_da_acq_no:,.2f}"
    ]

    summary_data['Note'] = [
        '',
        '',
        '',
        '',
        f'{n_pnrr/n_strutture*100:.1f}%',
        f'{n_non_pnrr/n_strutture*100:.1f}%',
        '',
        'Richiede nuovo finanziamento',
        'SCADENZA: MARZO 2026',
        '',
        '',
        'Budget già allocato',
        'Dotazioni operative',
        '',
        'Per completamento rete'
    ]

    df_summary = pd.DataFrame(summary_data)

    # 2. ANALISI PNRR PRIORITARIA
    print("  → Analisi PNRR Prioritaria")
    df_pnrr_detail = df_da_acq[df_da_acq['PNRR'] == 'SI'].groupby(
        ['Nome_Struttura', 'Zona', 'Tipologia']
    ).agg({
        'Costo_Totale': 'sum'
    }).reset_index().sort_values('Costo_Totale', ascending=False)

    df_pnrr_detail.columns = ['Struttura', 'Zona', 'Tipo', 'Fabbisogno (€)']

    # 3. FABBISOGNO PER DOTAZIONE
    print("  → Fabbisogno per Dotazione")
    df_fabb_dot = df_da_acq.groupby(['Categoria', 'Descrizione', 'Costo_Unitario_EUR']).agg({
        'Quantita_Da_Acquistare': 'sum',
        'Costo_Totale': 'sum'
    }).reset_index().sort_values(['Categoria', 'Costo_Totale'], ascending=[True, False])

    df_fabb_dot.columns = ['Categoria', 'Descrizione', 'Costo Unitario (€)', 'Quantità', 'Costo Totale (€)']

    # 4. DETTAGLIO PER STRUTTURA
    print("  → Dettaglio per Struttura")
    df_strutt_detail = df_merge.groupby(['Nome_Struttura', 'Zona', 'Tipologia', 'PNRR']).agg({
        'Costo_Totale': 'sum'
    }).reset_index().sort_values(['PNRR', 'Costo_Totale'], ascending=[False, False])

    df_strutt_detail.columns = ['Struttura', 'Zona', 'Tipo', 'PNRR', 'Fabbisogno (€)']

    # 5. CONFIGURAZIONI COMPLETE
    print("  → Configurazioni Complete")
    df_config = df_merge[[
        'Nome_Struttura', 'Zona', 'PNRR', 'Categoria', 'Descrizione',
        'Stato_Finanziamento', 'Quantita_Presente', 'Quantita_Richiesta',
        'Quantita_Da_Acquistare', 'Costo_Unitario_EUR', 'Costo_Totale'
    ]].sort_values(['Nome_Struttura', 'Categoria', 'Descrizione'])

    df_config.columns = [
        'Struttura', 'Zona', 'PNRR', 'Categoria', 'Dotazione',
        'Stato Finanziamento', 'Qty Presente', 'Qty Richiesta',
        'Qty Da Acquistare', 'Costo Unitario (€)', 'Costo Totale (€)'
    ]

    # Scrivi Excel
    print("  → Scrittura file Excel")
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
        df_pnrr_detail.to_excel(writer, sheet_name='PNRR Priorità', index=False)
        df_fabb_dot.to_excel(writer, sheet_name='Fabbisogno per Dotazione', index=False)
        df_strutt_detail.to_excel(writer, sheet_name='Fabbisogno per Struttura', index=False)
        df_config.to_excel(writer, sheet_name='Configurazioni Complete', index=False)

    # Formattazione
    print("  → Formattazione celle")
    wb = load_workbook(filename)

    # Formatta Executive Summary
    ws = wb['Executive Summary']
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30

    # Header
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Evidenzia totale
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if 'FABBISOGNO DA FINANZIARE' in str(row[0].value) or 'TOTALE INVESTIMENTO' in str(row[0].value):
            for cell in row:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Formatta altre schede
    for sheet_name in ['PNRR Priorità', 'Fabbisogno per Dotazione', 'Fabbisogno per Struttura']:
        ws = wb[sheet_name]

        # Header
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filename)

    print(f"\n✅ Report generato: {filename}")
    print(f"\n📊 RIEPILOGO:")
    print(f"  • Fabbisogno PNRR (priorità):  € {costo_da_acq_pnrr:,.2f}")
    print(f"  • Fabbisogno non-PNRR:         € {costo_da_acq_no:,.2f}")
    print(f"  • TOTALE DA FINANZIARE:        € {costo_da_acq_pnrr + costo_da_acq_no:,.2f}")
    print(f"\n  ⚠️  SCADENZA PNRR: MARZO 2026")

    return filename


if __name__ == "__main__":
    print("=" * 80)
    print("GENERAZIONE REPORT DIREZIONE - TELEMEDICINA USL TOSCANA NORD OVEST")
    print("=" * 80)
    print()

    filename = crea_report_direzione()

    print("\n" + "=" * 80)
    print("Il report è pronto per la presentazione in direzione!")
    print("=" * 80)
