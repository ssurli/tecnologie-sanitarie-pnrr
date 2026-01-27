#!/usr/bin/env python3
"""
Script per generare il Report Excel per Ordini e Acquisti
Tecnologie Sanitarie PNRR - USL Toscana Nord Ovest

Genera un file Excel con:
- Riepilogo generale
- CDC PNRR e non-PNRR
- ODC e Cure Intermedie PNRR e non-PNRR
- Fabbisogno complessivo per ordini
- Base per aggiornamenti

Autore: Claude Code
Data: 2026-01-27
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Percorso base
BASE_PATH = os.path.dirname(os.path.abspath(__file__))

# Stili
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
PNRR_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NON_PNRR_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
DA_ACQUISTARE_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def load_data():
    """Carica tutti i dati necessari dai file CSV."""
    print("Caricamento dati...")

    # Strutture sanitarie
    strutture = pd.read_csv(os.path.join(BASE_PATH, 'strutture_sanitarie.csv'))

    # Catalogo dotazioni
    catalogo = pd.read_csv(os.path.join(BASE_PATH, 'dotazioni_telemedicina_catalogo.csv'))

    # Dotazioni per struttura
    dotazioni = pd.read_csv(os.path.join(BASE_PATH, 'dotazioni_strutture_telemedicina.csv'))

    # Tecnologie arredi PNRR (se esiste)
    arredi_path = os.path.join(BASE_PATH, 'tecnologie_arredi_pnrr.csv')
    if os.path.exists(arredi_path):
        arredi = pd.read_csv(arredi_path)
    else:
        arredi = pd.DataFrame()

    return strutture, catalogo, dotazioni, arredi


def merge_data(strutture, catalogo, dotazioni):
    """Unisce i dati delle strutture, catalogo e dotazioni."""
    print("Elaborazione dati...")

    # Merge dotazioni con catalogo
    df = dotazioni.merge(
        catalogo[['Codice', 'Categoria', 'Descrizione', 'Costo_Unitario_EUR', 'IVA_Percentuale']].rename(
            columns={'Codice': 'Codice_Catalogo'}
        ),
        left_on='Codice_Dotazione',
        right_on='Codice_Catalogo',
        how='left'
    )

    # Rimuovi colonna duplicata
    if 'Codice_Catalogo' in df.columns:
        df = df.drop(columns=['Codice_Catalogo'])

    # Merge con strutture
    df = df.merge(
        strutture[['Codice', 'Tipologia', 'Nome_Struttura', 'Zona', 'Classificazione', 'PNRR', 'Posti_Letto']].rename(
            columns={'Codice': 'Codice_Str'}
        ),
        left_on='Codice_Struttura',
        right_on='Codice_Str',
        how='left'
    )

    # Rimuovi colonna duplicata
    if 'Codice_Str' in df.columns:
        df = df.drop(columns=['Codice_Str'])

    # Calcola quantità da acquistare
    df['Quantita_Da_Acquistare'] = (df['Quantita_Richiesta'] - df['Quantita_Presente']).clip(lower=0)

    # Calcola costo totale
    df['Costo_Totale'] = df['Quantita_Da_Acquistare'] * df['Costo_Unitario_EUR']

    # Rinomina colonne per chiarezza
    df = df.rename(columns={
        'Codice_Dotazione': 'Codice_Dispositivo'
    })

    return df


def auto_adjust_column_width(worksheet, df, start_col=1):
    """Auto-dimensiona le colonne in base al contenuto."""
    for idx, col in enumerate(df.columns):
        max_length = len(str(col))
        for cell in df[col].fillna('').astype(str):
            cell_len = len(str(cell))
            if cell_len > max_length:
                max_length = cell_len
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[get_column_letter(start_col + idx)].width = adjusted_width


def write_df_to_sheet(ws, df, start_row=1, with_header=True):
    """Scrive un DataFrame in un foglio Excel con formattazione."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=with_header)):
        for c_idx, value in enumerate(row):
            cell = ws.cell(row=start_row + r_idx, column=c_idx + 1, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=False, vertical='center')

            if r_idx == 0 and with_header:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    return start_row + len(df) + (1 if with_header else 0)


def create_riepilogo_sheet(wb, strutture, df):
    """Crea il foglio di riepilogo generale."""
    ws = wb.create_sheet("Riepilogo", 0)

    # Titolo
    ws['A1'] = "REPORT TECNOLOGIE SANITARIE - USL TOSCANA NORD OVEST"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells('A1:F1')

    ws['A2'] = f"Data generazione: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10)

    # Statistiche strutture
    row = 4
    ws.cell(row=row, column=1, value="STATISTICHE STRUTTURE").font = Font(bold=True, size=12)
    row += 1

    stats = [
        ("Totale Strutture", len(strutture)),
        ("Case della Comunità (CDC)", len(strutture[strutture['Tipologia'] == 'CdC'])),
        ("  - CDC Hub", len(strutture[(strutture['Tipologia'] == 'CdC') & (strutture['Classificazione'] == 'Hub')])),
        ("  - CDC Spoke", len(strutture[(strutture['Tipologia'] == 'CdC') & (strutture['Classificazione'] == 'Spoke')])),
        ("Ospedali di Comunità (ODC)", len(strutture[strutture['Tipologia'] == 'OdC'])),
        ("Cure Intermedie", len(strutture[strutture['Tipologia'] == 'Cure Intermedie'])),
        ("", ""),
        ("Strutture PNRR", len(strutture[strutture['PNRR'] == 'SI'])),
        ("Strutture non-PNRR", len(strutture[strutture['PNRR'] == 'NO'])),
    ]

    for label, value in stats:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if "PNRR" in label and value:
            if "non-PNRR" in label:
                ws.cell(row=row, column=2).fill = NON_PNRR_FILL
            else:
                ws.cell(row=row, column=2).fill = PNRR_FILL
        row += 1

    # Statistiche fabbisogno
    row += 2
    ws.cell(row=row, column=1, value="STATISTICHE FABBISOGNO").font = Font(bold=True, size=12)
    row += 1

    # Calcola statistiche
    da_acquistare = df[df['Stato_Finanziamento'] == 'DA_ACQUISTARE']
    finanziato = df[df['Stato_Finanziamento'] == 'FINANZIATO']
    presente = df[df['Stato_Finanziamento'] == 'PRESENTE']

    pnrr_da_acq = da_acquistare[da_acquistare['PNRR'] == 'SI']
    non_pnrr_da_acq = da_acquistare[da_acquistare['PNRR'] == 'NO']

    stats_fabb = [
        ("Totale dispositivi da acquistare", int(da_acquistare['Quantita_Da_Acquistare'].sum())),
        ("Costo totale da finanziare", f"€ {da_acquistare['Costo_Totale'].sum():,.2f}"),
        ("", ""),
        ("Dispositivi DA_ACQUISTARE PNRR", int(pnrr_da_acq['Quantita_Da_Acquistare'].sum())),
        ("Costo PNRR da acquistare", f"€ {pnrr_da_acq['Costo_Totale'].sum():,.2f}"),
        ("", ""),
        ("Dispositivi DA_ACQUISTARE non-PNRR", int(non_pnrr_da_acq['Quantita_Da_Acquistare'].sum())),
        ("Costo non-PNRR da acquistare", f"€ {non_pnrr_da_acq['Costo_Totale'].sum():,.2f}"),
        ("", ""),
        ("Dispositivi già finanziati/ordinati", int(finanziato['Quantita_Richiesta'].sum())),
        ("Dispositivi già presenti", int(presente['Quantita_Presente'].sum())),
    ]

    for label, value in stats_fabb:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Riepilogo per tipologia
    row += 2
    ws.cell(row=row, column=1, value="RIEPILOGO PER TIPOLOGIA STRUTTURA").font = Font(bold=True, size=12)
    row += 1

    # Header tabella
    headers = ["Tipologia", "N. Strutture", "Disp. Da Acquistare", "Costo Totale", "PNRR", "non-PNRR"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 1

    for tipologia in ['CdC', 'OdC', 'Cure Intermedie']:
        tip_strutture = strutture[strutture['Tipologia'] == tipologia]
        tip_df = da_acquistare[da_acquistare['Tipologia'] == tipologia]
        tip_pnrr = tip_df[tip_df['PNRR'] == 'SI']
        tip_non_pnrr = tip_df[tip_df['PNRR'] == 'NO']

        values = [
            tipologia,
            len(tip_strutture),
            int(tip_df['Quantita_Da_Acquistare'].sum()),
            tip_df['Costo_Totale'].sum(),
            tip_pnrr['Costo_Totale'].sum(),
            tip_non_pnrr['Costo_Totale'].sum()
        ]

        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = THIN_BORDER
            if c >= 4:
                cell.number_format = '€ #,##0.00'
        row += 1

    # Totale
    total_values = [
        "TOTALE",
        len(strutture),
        int(da_acquistare['Quantita_Da_Acquistare'].sum()),
        da_acquistare['Costo_Totale'].sum(),
        pnrr_da_acq['Costo_Totale'].sum(),
        non_pnrr_da_acq['Costo_Totale'].sum()
    ]
    for c, v in enumerate(total_values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER
        if c >= 4:
            cell.number_format = '€ #,##0.00'

    # Imposta larghezza colonne
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20


def create_dettaglio_sheet(wb, df, sheet_name, tipologie, pnrr_filter):
    """Crea un foglio con il dettaglio delle strutture."""
    ws = wb.create_sheet(sheet_name)

    # Filtra i dati
    mask = df['Tipologia'].isin(tipologie)
    if pnrr_filter == 'SI':
        mask = mask & (df['PNRR'] == 'SI')
    elif pnrr_filter == 'NO':
        mask = mask & (df['PNRR'] == 'NO')

    filtered_df = df[mask].copy()

    if filtered_df.empty:
        ws['A1'] = "Nessun dato disponibile per questa selezione"
        return

    # Prepara DataFrame per export
    export_df = filtered_df[[
        'Codice_Struttura', 'Nome_Struttura', 'Zona', 'Classificazione', 'PNRR',
        'Codice_Dispositivo', 'Categoria', 'Descrizione',
        'Quantita_Presente', 'Quantita_Richiesta', 'Quantita_Da_Acquistare',
        'Stato_Finanziamento', 'Costo_Unitario_EUR', 'Costo_Totale', 'Note'
    ]].sort_values(['Nome_Struttura', 'Descrizione'])

    # Rinomina colonne per chiarezza
    export_df.columns = [
        'Codice', 'Struttura', 'Zona', 'Classificazione', 'PNRR',
        'Cod. Dispositivo', 'Categoria', 'Descrizione Dispositivo',
        'Qtà Presente', 'Qtà Richiesta', 'Qtà Da Acquistare',
        'Stato', 'Costo Unitario €', 'Costo Totale €', 'Note'
    ]

    # Titolo
    titolo = f"Dettaglio {sheet_name.replace('_', ' ')}"
    ws['A1'] = titolo
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    # Scrivi dati
    write_df_to_sheet(ws, export_df, start_row=3)

    # Formattazione condizionale per stato
    for row in range(4, 4 + len(export_df)):
        stato_cell = ws.cell(row=row, column=12)
        if stato_cell.value == 'DA_ACQUISTARE':
            stato_cell.fill = DA_ACQUISTARE_FILL
        elif stato_cell.value == 'FINANZIATO':
            stato_cell.fill = PNRR_FILL

    # Riga totale
    total_row = 4 + len(export_df)
    ws.cell(row=total_row, column=1, value="TOTALE").font = TOTAL_FONT
    ws.cell(row=total_row, column=11, value=export_df['Qtà Da Acquistare'].sum()).font = TOTAL_FONT
    ws.cell(row=total_row, column=14, value=export_df['Costo Totale €'].sum())
    ws.cell(row=total_row, column=14).font = TOTAL_FONT
    ws.cell(row=total_row, column=14).number_format = '€ #,##0.00'

    for c in range(1, 16):
        ws.cell(row=total_row, column=c).fill = TOTAL_FILL
        ws.cell(row=total_row, column=c).border = THIN_BORDER

    auto_adjust_column_width(ws, export_df)


def create_fabbisogno_acquisti_sheet(wb, df):
    """Crea il foglio con il fabbisogno complessivo per gli ordini."""
    ws = wb.create_sheet("Fabbisogno_Acquisti")

    # Filtra solo DA_ACQUISTARE con quantità > 0
    da_acquistare = df[
        (df['Stato_Finanziamento'] == 'DA_ACQUISTARE') &
        (df['Quantita_Da_Acquistare'] > 0)
    ].copy()

    # Titolo
    ws['A1'] = "FABBISOGNO COMPLESSIVO PER ORDINI E ACQUISTI"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells('A1:H1')

    ws['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y')} - Solo dispositivi con stato DA_ACQUISTARE"
    ws['A2'].font = Font(italic=True, size=10)

    # SEZIONE 1: Aggregato per tipo dispositivo
    row = 4
    ws.cell(row=row, column=1, value="AGGREGATO PER TIPO DISPOSITIVO").font = Font(bold=True, size=12)
    row += 1

    # Aggrega per dispositivo
    agg_dispositivo = da_acquistare.groupby(['Codice_Dispositivo', 'Descrizione', 'Costo_Unitario_EUR']).agg({
        'Quantita_Da_Acquistare': 'sum',
        'Costo_Totale': 'sum'
    }).reset_index()

    agg_dispositivo = agg_dispositivo.rename(columns={
        'Codice_Dispositivo': 'Codice',
        'Descrizione': 'Descrizione Dispositivo',
        'Costo_Unitario_EUR': 'Costo Unitario €',
        'Quantita_Da_Acquistare': 'Quantità Totale',
        'Costo_Totale': 'Importo Totale €'
    }).sort_values('Importo Totale €', ascending=False)

    row = write_df_to_sheet(ws, agg_dispositivo, start_row=row)

    # Totale sezione 1
    ws.cell(row=row, column=1, value="TOTALE").font = TOTAL_FONT
    ws.cell(row=row, column=4, value=agg_dispositivo['Quantità Totale'].sum()).font = TOTAL_FONT
    ws.cell(row=row, column=5, value=agg_dispositivo['Importo Totale €'].sum())
    ws.cell(row=row, column=5).font = TOTAL_FONT
    ws.cell(row=row, column=5).number_format = '€ #,##0.00'
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER

    # SEZIONE 2: Dettaglio per struttura (per ordini)
    row += 3
    ws.cell(row=row, column=1, value="DETTAGLIO PER STRUTTURA (per ordini)").font = Font(bold=True, size=12)
    row += 1

    # Prepara dettaglio
    dettaglio = da_acquistare[[
        'Codice_Struttura', 'Nome_Struttura', 'Tipologia', 'PNRR', 'Zona',
        'Codice_Dispositivo', 'Descrizione',
        'Quantita_Da_Acquistare', 'Costo_Unitario_EUR', 'Costo_Totale'
    ]].copy()

    dettaglio.columns = [
        'Cod. Struttura', 'Struttura', 'Tipo', 'PNRR', 'Zona',
        'Cod. Dispositivo', 'Descrizione Dispositivo',
        'Qtà', 'Costo Unit. €', 'Costo Tot. €'
    ]

    dettaglio = dettaglio.sort_values(['PNRR', 'Tipo', 'Struttura', 'Descrizione Dispositivo'], ascending=[False, True, True, True])

    row = write_df_to_sheet(ws, dettaglio, start_row=row)

    # Totale sezione 2
    ws.cell(row=row, column=1, value="TOTALE").font = TOTAL_FONT
    ws.cell(row=row, column=8, value=dettaglio['Qtà'].sum()).font = TOTAL_FONT
    ws.cell(row=row, column=10, value=dettaglio['Costo Tot. €'].sum())
    ws.cell(row=row, column=10).font = TOTAL_FONT
    ws.cell(row=row, column=10).number_format = '€ #,##0.00'
    for c in range(1, 11):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER

    # SEZIONE 3: Riepilogo per PNRR/non-PNRR
    row += 3
    ws.cell(row=row, column=1, value="RIEPILOGO PNRR / non-PNRR").font = Font(bold=True, size=12)
    row += 1

    pnrr_summary = da_acquistare.groupby('PNRR').agg({
        'Quantita_Da_Acquistare': 'sum',
        'Costo_Totale': 'sum'
    }).reset_index()

    pnrr_summary.columns = ['Finanziamento', 'Qtà Totale', 'Importo Totale €']
    pnrr_summary['Finanziamento'] = pnrr_summary['Finanziamento'].replace({'SI': 'PNRR', 'NO': 'Non-PNRR'})

    row = write_df_to_sheet(ws, pnrr_summary, start_row=row)

    auto_adjust_column_width(ws, dettaglio)


def create_base_aggiornamenti_sheet(wb, strutture, catalogo, dotazioni):
    """Crea il foglio base per gli aggiornamenti futuri."""
    ws = wb.create_sheet("Base_Aggiornamenti")

    # Titolo
    ws['A1'] = "BASE DATI PER AGGIORNAMENTI"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells('A1:F1')

    ws['A2'] = "Utilizzare questo foglio per aggiornare le quantità e gli stati. Poi re-importare nel sistema."
    ws['A2'].font = Font(italic=True, size=10)
    ws['A3'] = "Campi editabili: Quantita_Presente, Quantita_Richiesta, Stato_Finanziamento, Note"
    ws['A3'].font = Font(italic=True, size=10, color="FF0000")

    # Prepara dati completi per aggiornamenti
    # Crea una riga per ogni combinazione struttura-dispositivo possibile

    update_df = dotazioni.copy()

    # Aggiungi info struttura
    update_df = update_df.merge(
        strutture[['Codice', 'Nome_Struttura', 'Tipologia', 'PNRR', 'Zona']],
        left_on='Codice_Struttura',
        right_on='Codice',
        how='left'
    )

    # Aggiungi info catalogo
    update_df = update_df.merge(
        catalogo[['Codice', 'Descrizione', 'Costo_Unitario_EUR']],
        left_on='Codice_Dotazione',
        right_on='Codice',
        how='left',
        suffixes=('_Str', '_Cat')
    )

    # Seleziona e ordina colonne
    export_df = update_df[[
        'Codice_Struttura', 'Nome_Struttura', 'Tipologia', 'PNRR', 'Zona',
        'Codice_Dotazione', 'Descrizione', 'Costo_Unitario_EUR',
        'Quantita_Presente', 'Quantita_Richiesta', 'Stato_Finanziamento', 'Note'
    ]].copy()

    export_df.columns = [
        'Codice_Struttura', 'Nome_Struttura', 'Tipologia', 'PNRR', 'Zona',
        'Codice_Dotazione', 'Descrizione_Dotazione', 'Costo_Unitario',
        'Quantita_Presente', 'Quantita_Richiesta', 'Stato_Finanziamento', 'Note'
    ]

    export_df = export_df.sort_values(['Tipologia', 'Nome_Struttura', 'Codice_Dotazione'])

    # Scrivi dati
    write_df_to_sheet(ws, export_df, start_row=5)

    # Evidenzia colonne editabili
    for row in range(6, 6 + len(export_df)):
        for col in [9, 10, 11, 12]:  # Colonne editabili
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    auto_adjust_column_width(ws, export_df)

    # Aggiungi istruzioni
    row = 6 + len(export_df) + 2
    ws.cell(row=row, column=1, value="ISTRUZIONI PER L'AGGIORNAMENTO:").font = Font(bold=True)
    row += 1
    istruzioni = [
        "1. Modificare i valori nelle colonne evidenziate in giallo",
        "2. Stati validi: DA_ACQUISTARE, FINANZIATO, PRESENTE, NON_RICHIESTO",
        "3. Salvare il file con un nuovo nome (es: aggiornamento_DDMMYYYY.xlsx)",
        "4. Eseguire lo script di importazione per aggiornare il database",
        "5. Verificare i dati sulla dashboard Streamlit"
    ]
    for istr in istruzioni:
        ws.cell(row=row, column=1, value=istr)
        row += 1


def create_arredi_pnrr_sheet(wb, arredi):
    """Crea il foglio con gli arredi PNRR se disponibili."""
    if arredi.empty:
        return

    ws = wb.create_sheet("Arredi_PNRR")

    # Titolo
    ws['A1'] = "ARREDI E TECNOLOGIE PNRR"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells('A1:F1')

    ws['A2'] = "Dettaglio arredi e tecnologie per strutture PNRR"
    ws['A2'].font = Font(italic=True, size=10)

    # Scrivi dati
    write_df_to_sheet(ws, arredi, start_row=4)

    # Totale
    total_row = 5 + len(arredi)
    ws.cell(row=total_row, column=1, value="TOTALE").font = TOTAL_FONT

    if 'Totale' in arredi.columns:
        totale = arredi['Totale'].sum()
    elif 'Quantita' in arredi.columns:
        totale = arredi['Quantita'].sum()
    else:
        totale = 0

    ws.cell(row=total_row, column=7, value=totale)
    ws.cell(row=total_row, column=7).font = TOTAL_FONT
    ws.cell(row=total_row, column=7).number_format = '€ #,##0.00'

    for c in range(1, 8):
        ws.cell(row=total_row, column=c).fill = TOTAL_FILL
        ws.cell(row=total_row, column=c).border = THIN_BORDER

    auto_adjust_column_width(ws, arredi)


def main():
    """Funzione principale per generare il report Excel."""
    print("=" * 60)
    print("GENERAZIONE REPORT EXCEL - TECNOLOGIE SANITARIE PNRR")
    print("=" * 60)

    # Carica dati
    strutture, catalogo, dotazioni, arredi = load_data()

    # Elabora dati
    df = merge_data(strutture, catalogo, dotazioni)

    # Crea workbook
    wb = Workbook()

    # Rimuovi il foglio di default
    wb.remove(wb.active)

    print("Creazione fogli Excel...")

    # 1. Foglio Riepilogo
    print("  - Riepilogo generale")
    create_riepilogo_sheet(wb, strutture, df)

    # 2. CDC PNRR
    print("  - CDC PNRR")
    create_dettaglio_sheet(wb, df, "CDC_PNRR", ['CdC'], 'SI')

    # 3. CDC non-PNRR
    print("  - CDC non-PNRR")
    create_dettaglio_sheet(wb, df, "CDC_non_PNRR", ['CdC'], 'NO')

    # 4. ODC e Cure Intermedie PNRR
    print("  - ODC e Cure Intermedie PNRR")
    create_dettaglio_sheet(wb, df, "ODC_CureInt_PNRR", ['OdC', 'Cure Intermedie'], 'SI')

    # 5. ODC e Cure Intermedie non-PNRR
    print("  - ODC e Cure Intermedie non-PNRR")
    create_dettaglio_sheet(wb, df, "ODC_CureInt_non_PNRR", ['OdC', 'Cure Intermedie'], 'NO')

    # 6. Fabbisogno Acquisti
    print("  - Fabbisogno per Acquisti")
    create_fabbisogno_acquisti_sheet(wb, df)

    # 7. Base Aggiornamenti
    print("  - Base per Aggiornamenti")
    create_base_aggiornamenti_sheet(wb, strutture, catalogo, dotazioni)

    # 8. Arredi PNRR (se disponibili)
    if not arredi.empty:
        print("  - Arredi PNRR")
        create_arredi_pnrr_sheet(wb, arredi)

    # Salva file
    output_filename = f"Report_Tecnologie_Sanitarie_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = os.path.join(BASE_PATH, output_filename)

    print(f"\nSalvataggio file: {output_filename}")
    wb.save(output_path)

    print("\n" + "=" * 60)
    print(f"Report generato con successo!")
    print(f"File: {output_path}")
    print("=" * 60)

    # Statistiche finali
    da_acq = df[df['Stato_Finanziamento'] == 'DA_ACQUISTARE']
    print(f"\nStatistiche:")
    print(f"  - Dispositivi da acquistare: {int(da_acq['Quantita_Da_Acquistare'].sum())}")
    print(f"  - Costo totale: € {da_acq['Costo_Totale'].sum():,.2f}")

    return output_path


if __name__ == "__main__":
    main()
